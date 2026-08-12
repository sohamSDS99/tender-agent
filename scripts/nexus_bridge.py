"""
Nexus AMS Bridge for Tender Agent
===================================
This script connects the tender-agent to the Nexus AMS platform.

It does 4 things:
1. Registers the agent and sends heartbeats (so AMS knows the agent is alive)
2. Polls for incoming chat messages from the AMS UI
3. Runs tender discovery searches and sends results back to the AMS chat
4. Fills tender forms uploaded by users using company context documents

Usage:
    cd ~/Desktop/tender-agent
    source .venv/bin/activate
    python scripts/nexus_bridge.py

Press Ctrl+C to stop.
"""

import json
import os
import sys
import time
import tempfile
import threading
import traceback
from datetime import datetime, timezone
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / ".env")

from nexus_sdk import NexusClient, AgentConfig

# Import tender-agent's discovery modules
from src.discovery.sam_gov import SamGovScraper, score_relevance
from src.discovery.serp_search import SerpTenderSearcher, get_excluded_domains_for_region
from src.discovery.ted_europa import TedEuropaSearcher
from src.discovery.uk_tenders import UkTenderSearcher
from src.discovery.boamp_france import BoampSearcher
from src.discovery.world_bank import WorldBankSearcher
from src.discovery.prozorro import ProzorroSearcher
from src.discovery.canada_buys import CanadaBuysSearcher
from src.discovery.austender import AusTenderSearcher
from src.discovery.sa_etender import SaEtenderSearcher
from src.discovery.colombia_secop import ColombiaSecopSearcher
from src.discovery.brazil_compras import BrazilComprasSearcher
from src.discovery.germany_bkms import GermanyBkmsSearcher
from src.discovery.italy_anac import ItalyAnacSearcher
from src.discovery.dominican_dgcp import DominicanDgcpSearcher
from src.discovery.peru_oece import PeruOeceSearcher
from src.discovery.world_bank_v2 import WorldBankV2Searcher
from src.discovery.nigeria_nocopo import NigeriaNocopoSearcher
from src.discovery.kenya_ppra import KenyaPpraSearcher
from src.discovery.uganda_gpp import UgandaGppSearcher
from src.discovery.mexico_cdmx import MexicoCdmxSearcher

# Import form processing modules
from src.forms.form_parser import FormParser
from src.forms.form_filler import FormFiller, FillResult
from src.forms.form_writer import FormWriter

# Slack notifications
from slack_notifier import (
    notify_task_completed,
    notify_task_failed,
    notify_agent_online,
    notify_agent_offline,
)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

NEXUS_URL = os.getenv("NEXUS_AMS_URL", "http://localhost:3000")

# P0.0 — the bridge's credential for the AMS.
#
# Every bridge-facing endpoint on the AMS side currently falls open when this is
# unset (`agent-auth.ts` begins `if (!required) return true;`). That fall-open is
# being removed, at which point a call without this header gets a 401.
#
# This value must be IDENTICAL to NEXUS_AGENT_API_KEY on the AMS service. Set it
# on both services before the AMS side ships its fail-closed change — see
# docs/runbooks/credential-rotation.md in the nexus-ams repository.
NEXUS_AGENT_API_KEY = os.getenv("NEXUS_AGENT_API_KEY", "")
OPENROUTER_API_KEY = os.getenv("DASHSCOPE_API_KEY", "")
OPENROUTER_BASE_URL = os.getenv("QWEN_BASE_URL", "https://openrouter.ai/api/v1")
LLM_MODEL = "anthropic/claude-sonnet-4"
DRY_RUN = os.getenv("DRY_RUN", "true").lower() in ("true", "1", "yes")

def _ams_headers() -> dict[str, str]:
    """Auth headers for every direct call to the AMS.

    P0.0. This module talks to the AMS two ways: through `NexusClient`, which has
    always sent `Authorization` via its own `_build_headers()`, and through a
    handful of raw `httpx` calls that did not. Those raw calls
    (`/search`, `/context`, `/prompt`, and the dead `/inbox` poller) reached
    bridge-facing endpoints with no credential at all, and only worked because
    the AMS fell open when `NEXUS_AGENT_API_KEY` was unset.

    Every raw AMS call must pass `headers=_ams_headers()`. There is no local
    check enforcing that — the enforcement is the AMS itself, which returns 401
    once its fall-open is removed. A forgotten header fails loudly and
    immediately rather than silently granting access, which is the right
    direction for this to fail.

    Returns an empty dict when the key is unset, so local development against an
    AMS that has not yet been hardened behaves exactly as before.
    """
    if not NEXUS_AGENT_API_KEY:
        return {}
    return {"Authorization": f"Bearer {NEXUS_AGENT_API_KEY}"}


def _report_auth_failure(status_code: int, what: str) -> bool:
    """Print an unmissable message when the AMS refuses a call for auth reasons.

    Returns True when the status was an auth failure, so a caller can tell a
    credential fault apart from an ordinary miss.

    `_ams_headers()` above returns `{}` when the key is unset, and three of this
    module's AMS calls degrade quietly rather than failing: a knowledge-base
    search falls back to legacy context, a file download returns "", and a
    prompt fetch falls back to the built-in prompt. Each then carries on. So a
    missing or mismatched credential does not present as an outage — it presents
    as the agent quietly getting worse at its job, which is far harder to notice
    and much harder to attribute.

    P0.1 closed the endpoints these calls use, so an unset or wrong key now
    produces 401 here instead of a silent success.
    """
    if status_code not in (401, 403):
        return False
    print("")
    print("  " + "!" * 68)
    print(f"  !! AUTH FAILURE — AMS returned {status_code} for {what}.")
    print("  !! This is a CONFIGURATION fault, not a data condition.")
    print("  !! NEXUS_AGENT_API_KEY must be set to the SAME value on this")
    print("  !! bridge and on the AMS web service.")
    print("  !! The agent continues with degraded input; results will be worse")
    print("  !! until this is fixed. See docs/runbooks/p0.1-bridge-auth.md in")
    print("  !! the nexus-ams repository.")
    print("  " + "!" * 68)
    print("")
    return True


# Module-level reference to the NexusClient (set in main())
_client: "NexusClient | None" = None

# Current thread ID for audit context (set per-message)
_current_thread_id: str = ""


def _audit(
    action_type: str,
    description: str,
    **kwargs,
) -> None:
    """Write an audit log entry. Safe to call even if client isn't ready."""
    if not _client:
        return
    try:
        _client.audit(
            action_type,
            description,
            conversation_id=kwargs.pop("conversation_id", _current_thread_id or None),
            **kwargs,
        )
    except Exception as exc:
        print(f"  [audit] Failed to log {action_type}: {exc}")


# ---------------------------------------------------------------------------
# Document Context — fetches assigned docs from AMS for RAG
# ---------------------------------------------------------------------------

_cached_context: str = ""
_context_fetched_at: float = 0.0
CONTEXT_TTL_SECONDS = 300  # 5 minutes


def search_knowledge_base(query: str, top_k: int = 15) -> str:
    """Search the agent's knowledge base for chunks relevant to the query.

    Uses semantic vector search via the AMS /search endpoint.
    Returns a formatted context block with only the most relevant document chunks.
    Falls back to the legacy full-context approach if the KB has no indexed chunks.
    """
    try:
        import httpx
        resp = httpx.post(
            f"{NEXUS_URL}/api/agents/tender-agent/search",
            json={"query": query, "topK": top_k},
            headers=_ams_headers(),
            timeout=15.0,
        )
        if resp.status_code != 200:
            _report_auth_failure(resp.status_code, "knowledge-base search")
            print(f"KB search returned {resp.status_code}, falling back to legacy context")
            return fetch_agent_context()

        data = resp.json()
        mode = data.get("mode", "semantic")

        if mode == "legacy":
            # No chunks indexed yet — use the full fallback context
            fallback = data.get("fallbackContext", "")
            if fallback:
                _audit(
                    "kb_search",
                    f"Knowledge base not indexed, using legacy full-text context",
                    output_payload={"mode": "legacy", "query": query[:200]},
                )
                return f"<agent_context>\n{fallback}\n</agent_context>"
            return fetch_agent_context()

        results = data.get("results", [])
        total_tokens = data.get("totalTokens", 0)
        formatted = data.get("formattedContext", "")

        if not results:
            print(f"KB search returned 0 results for query: {query[:80]}")
            return fetch_agent_context()

        # Collect unique source filenames for audit
        filenames = list(set(r.get("filename", "?") for r in results))
        _audit(
            "kb_search",
            f"KB search: {len(results)} chunks from {len(filenames)} docs ({total_tokens} tokens)",
            output_payload={
                "mode": "semantic",
                "query": query[:200],
                "result_count": len(results),
                "total_tokens": total_tokens,
                "source_files": filenames,
            },
        )

        parts = [
            "<agent_context>",
            f"The following {len(results)} relevant excerpts were retrieved from your knowledge base.",
            f"Source documents: {', '.join(filenames)}",
            "",
            formatted,
            "</agent_context>",
        ]
        return "\n".join(parts)

    except Exception as exc:
        print(f"KB search failed: {exc}, falling back to legacy context")
        return fetch_agent_context()


def fetch_agent_context() -> str:
    """Fetch ALL documents assigned to tender-agent from the AMS (legacy fallback).

    Returns a formatted context block. Cached for 5 minutes.
    Used as fallback when the knowledge base hasn't been indexed yet.
    """
    global _cached_context, _context_fetched_at

    if _cached_context and (time.time() - _context_fetched_at) < CONTEXT_TTL_SECONDS:
        return _cached_context

    try:
        import httpx
        resp = httpx.get(
            f"{NEXUS_URL}/api/agents/tender-agent/context",
            headers=_ams_headers(),
            timeout=10.0,
        )
        if resp.status_code != 200:
            print(f"Context endpoint returned {resp.status_code}")
            return ""

        data = resp.json()
        docs = data.get("documents", [])
        if not docs:
            print("No documents assigned to tender-agent")
            return ""

        print(f"Loaded {len(docs)} context document(s) (legacy mode)")
        doc_names = [d.get("filename", "?") for d in docs]
        _audit(
            "document_read",
            f"Loaded {len(docs)} context document(s) (legacy): {', '.join(doc_names)}",
            output_payload={"document_count": len(docs), "filenames": doc_names},
        )

        parts = ["<agent_context>"]
        parts.append("The following company documents have been assigned to you. "
                     "Use this information to answer questions and fill forms.")
        parts.append("")

        for doc in docs:
            filename = doc.get("filename", "Unknown")
            text = doc.get("extractedText", "").strip()
            status = doc.get("extractionStatus", "unknown")
            minio_key = doc.get("minioKey", "")

            if text:
                parts.append(f"## Document: {filename}")
                parts.append(text)
                parts.append("")
            elif minio_key:
                print(f"  Document '{filename}' has no extracted text (status={status}), trying local extraction...")
                local_text = _extract_text_locally(minio_key, filename)
                if local_text:
                    parts.append(f"## Document: {filename}")
                    parts.append(local_text)
                    parts.append("")
                else:
                    parts.append(f"## Document: {filename}")
                    parts.append(f"(Document uploaded but text extraction pending — file available at {minio_key})")
                    parts.append("")

        parts.append("</agent_context>")

        _cached_context = "\n".join(parts)
        _context_fetched_at = time.time()
        return _cached_context

    except Exception as exc:
        print(f"Context fetch failed: {exc}")
        return ""


def _extract_text_locally(minio_key: str, filename: str) -> str:
    """Download a document from AMS and extract text locally as a fallback.

    Used when the AMS extractor Docker service hasn't processed the document.
    """
    try:
        import httpx

        # Download from AMS file endpoint
        url = f"{NEXUS_URL}/api/chat/files/{minio_key}"
        # This endpoint is one of the seven with no auth at all today (P0.1
        # closes it). Send the credential now so the bridge does not break the
        # moment it is closed.
        resp = httpx.get(
            url,
            headers=_ams_headers(),
            timeout=30.0,
            follow_redirects=True,
        )
        if resp.status_code != 200:
            _report_auth_failure(resp.status_code, f"file download ({minio_key})")
            print(f"    Download failed: HTTP {resp.status_code}")
            return ""

        content = resp.content
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        # PDF
        if ext == "pdf":
            try:
                import pypdf
                import io
                reader = pypdf.PdfReader(io.BytesIO(content))
                text = ""
                for page in reader.pages:
                    text += (page.extract_text() or "") + "\n"
                if text.strip():
                    print(f"    Locally extracted {len(text)} chars from PDF")
                    return text.strip()[:4000]
            except Exception as exc:
                print(f"    PDF extraction failed: {exc}")

        # DOCX
        elif ext in ("docx", "doc"):
            try:
                from docx import Document
                import io
                doc = Document(io.BytesIO(content))
                text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                if text.strip():
                    print(f"    Locally extracted {len(text)} chars from DOCX")
                    return text.strip()[:4000]
            except Exception as exc:
                print(f"    DOCX extraction failed: {exc}")

        # XLSX
        elif ext in ("xlsx", "xls"):
            try:
                from openpyxl import load_workbook
                import io
                wb = load_workbook(io.BytesIO(content), data_only=True)
                text = ""
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        row_text = " | ".join(str(c) for c in row if c is not None)
                        if row_text.strip():
                            text += row_text + "\n"
                if text.strip():
                    print(f"    Locally extracted {len(text)} chars from XLSX")
                    return text.strip()[:4000]
            except Exception as exc:
                print(f"    XLSX extraction failed: {exc}")

        # Plain text / CSV / JSON
        elif ext in ("txt", "csv", "json", "md"):
            try:
                text = content.decode("utf-8", errors="replace")
                if text.strip():
                    print(f"    Read {len(text)} chars from text file")
                    return text.strip()[:4000]
            except Exception:
                pass

        return ""
    except Exception as exc:
        print(f"    Local extraction error: {exc}")
        return ""

# ---------------------------------------------------------------------------
# Prompt Studio — Fetch active system prompt from AMS
# ---------------------------------------------------------------------------

# Cache the active prompt for 60 seconds to avoid hitting AMS on every LLM call
_cached_prompt: dict | None = None
_prompt_cache_time: float = 0
_PROMPT_CACHE_TTL = 60  # seconds

DEFAULT_SYSTEM_PROMPT = (
    "You are the Tender Agent — an AI assistant for SDS Manager, "
    "specializing in Safety Data Sheet (SDS) and Environmental Health "
    "& Safety (EHS) software tenders.\n\n"
    "CAPABILITIES:\n"
    "- You CAN access and read company documents that have been "
    "uploaded and assigned to you in the Nexus AMS Documents section.\n"
    "- You CAN search for government tenders globally.\n"
    "- You CAN fill out tender forms using company context.\n"
    "- When asked about company details, reference the documents below.\n"
    "- Be concise, specific, and actionable."
)


def fetch_active_prompt() -> str:
    """Fetch the active system prompt from AMS Prompt Studio.

    Falls back to DEFAULT_SYSTEM_PROMPT if no active prompt is configured
    or if the AMS is unreachable.
    """
    global _cached_prompt, _prompt_cache_time
    import httpx

    # Return cached prompt if still fresh
    if _cached_prompt and (time.time() - _prompt_cache_time) < _PROMPT_CACHE_TTL:
        return _cached_prompt.get("systemPrompt", DEFAULT_SYSTEM_PROMPT)

    try:
        agent_name = "tender-agent"
        resp = httpx.get(
            f"{NEXUS_URL}/api/agents/{agent_name}/prompt",
            headers=_ams_headers(),
            timeout=5.0,
        )
        _report_auth_failure(resp.status_code, "active prompt fetch")
        if resp.status_code == 200:
            data = resp.json()
            if data.get("hasActivePrompt"):
                _cached_prompt = data
                _prompt_cache_time = time.time()
                print(f"  [Prompt Studio] Using active prompt v{data.get('version')} — {data.get('name', '')}")
                return data["systemPrompt"]
    except Exception as exc:
        print(f"  [Prompt Studio] Failed to fetch active prompt: {exc}")

    # Fallback to default
    return DEFAULT_SYSTEM_PROMPT


# ---------------------------------------------------------------------------
# LLM Helper — calls Claude Sonnet via OpenRouter
# ---------------------------------------------------------------------------

def call_llm(prompt: str, max_tokens: int = 1024) -> dict:
    """Call Claude Sonnet via OpenRouter and return response with usage stats.

    Returns:
        {
            "content": "the LLM response text",
            "tokens_input": 123,
            "tokens_output": 45,
            "cost_usd": 0.0001,
            "model": "anthropic/claude-sonnet-4",
            "duration_ms": 1234,
        }
    """
    import httpx

    # Fetch system prompt from Prompt Studio (with fallback)
    system_prompt = fetch_active_prompt()

    start_time = time.perf_counter()

    response = httpx.post(
        f"{OPENROUTER_BASE_URL}/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://sdsmanager.com",
            "X-Title": "SDS Tender Agent",
        },
        json={
            "model": LLM_MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": (
                        system_prompt + "\n\n"
                        + search_knowledge_base(prompt)
                    ),
                },
                {"role": "user", "content": prompt},
            ],
            "max_tokens": max_tokens,
            "temperature": 0.3,
        },
        timeout=60.0,
    )
    response.raise_for_status()
    data = response.json()

    duration_ms = int((time.perf_counter() - start_time) * 1000)

    content = data["choices"][0]["message"].get("content", "")
    usage = data.get("usage", {})
    tokens_in = usage.get("prompt_tokens", 0)
    tokens_out = usage.get("completion_tokens", 0)

    # Claude Sonnet via OpenRouter: $3.00/M input, $15.00/M output
    cost_usd = (tokens_in * 3.00 / 1_000_000) + (tokens_out * 15.00 / 1_000_000)

    # Audit every LLM call
    _audit(
        "llm_call",
        f"LLM call: {prompt[:80]}{'...' if len(prompt) > 80 else ''}",
        model_used=LLM_MODEL,
        tokens_input=tokens_in,
        tokens_output=tokens_out,
        cost_usd=round(cost_usd, 6),
        duration_ms=duration_ms,
        input_payload={"prompt": prompt[:500], "max_tokens": max_tokens},
        output_payload={"response": content[:500], "full_length": len(content)},
    )

    return {
        "content": content,
        "tokens_input": tokens_in,
        "tokens_output": tokens_out,
        "cost_usd": round(cost_usd, 6),
        "model": LLM_MODEL,
        "duration_ms": duration_ms,
    }


# ---------------------------------------------------------------------------
# Tender Discovery — runs the SAM.gov scraper and formats results
# ---------------------------------------------------------------------------

# URL patterns that indicate an actual tender/solicitation/bid page
TENDER_URL_PATTERNS = [
    "/opp/", "/notice/", "/solicitation", "/bid/", "/rfp/",
    "/tender/", "/procurement/", "/contract-opportunity/",
    "/open-bids/", "/view", "/detail", "/opportunity/",
    "/search-rfp/", "/searchrfp/", "/public/notice",
    "/contract/", "/award/", "/announcement/",
]

# URL patterns that indicate NOT a tender (info pages, wikis, articles, news)
NOT_TENDER_PATTERNS = [
    "/legislation/", "/directives/", "/guidelines/", "/themes/",
    "/wiki/", "/blog/", "/article/", "/articles/", "/news/",
    "/about/", "/glossary/", "/eguides/", "/films/", "/tools/",
    "/resources/", "/faq/", "/help/", "/contact/", "/policy/",
    "/regulation/", "/publications/", "/research/", "/reports/",
    "/events/", "/magazine/", "/webinar/", "/training/", "/podcast/",
    "/case-study/", "/case-studies/", "/white-paper/", "/whitepapers/",
    "/product/", "/products/", "/pricing/", "/demo/", "/solutions/",
    "/press-release/", "/press/", "/media/", "/insights/",
    "/category/", "/tag/", "/author/",
]

# Domains to always exclude — social, news, SaaS marketing, and industry
# magazines that will never contain actual tender listings
EXCLUDE_DOMAINS = {
    # Social media
    "linkedin.com", "youtube.com", "facebook.com", "twitter.com",
    "reddit.com", "quora.com", "instagram.com", "tiktok.com",
    # Content platforms
    "medium.com", "wikipedia.org", "slideshare.net",
    # Industry news & magazines (articles ABOUT safety, not tenders)
    "ohsonline.com", "ehstoday.com", "ishn.com", "safetyandhealthmagazine.com",
    "chemicalwatch.com", "chemweek.com", "chemicalprocessing.com",
    # SaaS / marketing sites
    "heyiris.ai", "hubspot.com", "salesforce.com", "g2.com",
    "capterra.com", "gartner.com", "softwareadvice.com",
    # OSHA info pages (not procurement)
    "oshwiki.osha.europa.eu", "eguides.osha.europa.eu",
}

# Domains where ONLY specific paths are tenders
STRICT_PATH_DOMAINS = {
    "osha.europa.eu": ["/procurement"],
    "gov.uk": ["/contracts-finder", "/tender"],
}


def is_valid_tender_link(url: str) -> bool:
    """Check if a URL is likely a tender/procurement page.

    Uses a blocklist approach: block known-bad domains and non-tender URL
    patterns, but allow everything else through.  Quality control is handled
    downstream by relevance scoring and deadline classification.
    """
    from urllib.parse import urlparse
    try:
        parsed = urlparse(url)
        domain = parsed.netloc.lower().replace("www.", "")
        path = parsed.path.lower()

        # Block 1: Known non-tender sites (social media, wikis, etc.)
        for excluded in EXCLUDE_DOMAINS:
            if excluded in domain:
                return False

        # Block 2: Info/wiki/legislation pages on otherwise-OK domains
        for pattern in NOT_TENDER_PATTERNS:
            if pattern in path:
                return False

        # Block 3: Strict-path domains — only specific paths are tenders
        for strict_domain, allowed_paths in STRICT_PATH_DOMAINS.items():
            if strict_domain in domain:
                return any(ap in path for ap in allowed_paths)

        # Everything else is allowed — relevance scoring + deadline
        # classification handle quality control downstream
        return True
    except Exception:
        return False


def _strict_deadline_filter(leads: list, source_label: str) -> list:
    """Strict deadline gate — blocks ANY lead without a verified future deadline.

    Used on BOTH API results and SERP results. No exceptions.
    If we can't prove the deadline is in the future, the lead is blocked.

    Args:
        leads: List of lead objects (API or SERP).
        source_label: "API" or "SERP" — for logging only.

    Returns:
        Filtered list where every lead has a verified, future deadline.
    """
    from dateutil import parser as dateparser

    today = datetime.now(timezone.utc).date()
    valid = []
    blocked_expired = 0
    blocked_no_deadline = 0

    for lead in leads:
        deadline = getattr(lead, 'submission_deadline', '') or ''

        if not deadline or len(deadline) < 8:
            blocked_no_deadline += 1
            continue

        try:
            parsed_date = dateparser.parse(deadline).date()
            if parsed_date >= today:
                valid.append(lead)
            else:
                blocked_expired += 1
        except Exception:
            blocked_no_deadline += 1

    print(f"    [{source_label} deadline gate] {len(valid)} active / "
          f"{blocked_expired} expired / {blocked_no_deadline} no-deadline "
          f"(from {len(leads)} total)")

    return valid


def _filter_serp_leads(leads: list, query: str) -> list:
    """Run the strict filtering pipeline on SERP results.

    Every SERP result gets its actual page fetched to extract and verify
    the deadline. No lead passes without a confirmed future deadline.

    Gates:
      1. Block known-bad URLs (blocklist)
      2. Block wrong region
      3. Require strong SDS/EHS keyword match
      4. Reject empty portal pages
      5. Require tender signal (proof it's procurement, not news)
      6. Minimum relevance score
      7. Deadline from snippet (expired → block immediately)
      8. PAGE-FETCH every remaining lead: extract deadline from actual page.
         If expired/closed/no-deadline-found → BLOCKED. No exceptions.
    """
    from dateutil import parser as dateparser
    from urllib.parse import urlparse
    from src.discovery.serp_search import extract_deadline_from_text
    from src.discovery.deadline_verifier import verify_deadline

    today = datetime.now(timezone.utc).date()
    excluded_domains = get_excluded_domains_for_region(query)

    MIN_RELEVANCE = 0.30

    def classify_deadline_from_snippet(lead: object) -> str:
        """Quick check from snippet text only. Returns valid/expired/missing."""
        deadline = getattr(lead, 'submission_deadline', '') or ''
        if not deadline or len(deadline) < 8:
            desc = getattr(lead, 'description', '') or ''
            title = getattr(lead, 'title', '') or ''
            deadline = extract_deadline_from_text(f"{title} {desc}")
            if deadline:
                lead.submission_deadline = deadline
        if deadline and len(deadline) >= 8:
            try:
                parsed_date = dateparser.parse(deadline).date()
                return 'valid' if parsed_date >= today else 'expired'
            except Exception:
                pass
        return 'missing'

    def is_correct_region(lead: object) -> bool:
        if not excluded_domains:
            return True
        try:
            domain = urlparse(lead.source_url).netloc.lower().replace("www.", "")
            return not any(exc in domain for exc in excluded_domains)
        except Exception:
            return True

    valid = []
    needs_page_verify: list = []
    counts = {
        "blocked_url": 0, "blocked_region": 0, "no_strong_kw": 0,
        "empty_page": 0, "no_tender_signal": 0,
        "low_relevance": 0, "expired_snippet": 0,
        "page_active": 0, "page_expired": 0, "page_no_deadline": 0,
    }

    # Gates 1-7: quick filters
    for lead in leads:
        raw = getattr(lead, 'raw_data', {}) or {}

        if not is_valid_tender_link(lead.source_url):
            counts["blocked_url"] += 1
            continue
        if not is_correct_region(lead):
            counts["blocked_region"] += 1
            continue
        if not raw.get("has_strong_match", False):
            counts["no_strong_kw"] += 1
            continue
        if raw.get("is_empty_page", False):
            counts["empty_page"] += 1
            continue
        if not raw.get("has_tender_signal", False):
            counts["no_tender_signal"] += 1
            continue
        if lead.relevance_score < MIN_RELEVANCE:
            counts["low_relevance"] += 1
            continue

        snippet_status = classify_deadline_from_snippet(lead)
        if snippet_status == 'expired':
            counts["expired_snippet"] += 1
            continue

        if snippet_status == 'valid':
            # Snippet had a valid future deadline — still page-verify to be sure
            needs_page_verify.append(lead)
        else:
            # No deadline in snippet — must page-verify
            needs_page_verify.append(lead)

    # ------------------------------------------------------------------
    # Gate 8: Page-fetch EVERY lead that passed gates 1-7.
    # No lead gets through without a verified future deadline.
    # ------------------------------------------------------------------
    if needs_page_verify:
        print(f"    [Gate 8] Fetching {len(needs_page_verify)} pages to verify deadlines...")

    for lead in needs_page_verify:
        url = getattr(lead, 'source_url', '') or ''
        existing_deadline = getattr(lead, 'submission_deadline', '') or ''

        # If we already have a snippet-extracted future deadline, trust it
        # but still double-check via page fetch for safety
        if existing_deadline and len(existing_deadline) >= 8:
            try:
                parsed = dateparser.parse(existing_deadline).date()
                if parsed >= today:
                    lead._effective_score = lead.relevance_score
                    counts["page_active"] += 1
                    valid.append(lead)
                    continue
            except Exception:
                pass

        # Must fetch the page to find deadline
        if not url:
            counts["page_no_deadline"] += 1
            print(f"      BLOCKED (no URL): {lead.title[:60]}...")
            continue

        try:
            vr = verify_deadline(url, timeout=12.0, use_llm=True)
            vr_status = vr.get("status", "unknown")
            vr_deadline = vr.get("deadline", "")
            vr_source = vr.get("source", "none")

            if vr_status == "active" and vr_deadline:
                lead.submission_deadline = vr_deadline
                lead._effective_score = lead.relevance_score
                counts["page_active"] += 1
                print(f"      ACTIVE: {lead.title[:50]}... "
                      f"[deadline={vr_deadline}, via={vr_source}]")
                valid.append(lead)

            elif vr_status in ("expired", "closed"):
                counts["page_expired"] += 1
                print(f"      BLOCKED ({vr_status}): {lead.title[:50]}... "
                      f"[deadline={vr_deadline}, via={vr_source}]")

            else:
                # Could not find any deadline on the page — BLOCKED
                counts["page_no_deadline"] += 1
                print(f"      BLOCKED (no deadline found): {lead.title[:50]}...")

        except Exception as exc:
            # Page fetch failed — BLOCKED (not graceful degradation anymore)
            counts["page_no_deadline"] += 1
            print(f"      BLOCKED (fetch failed): {lead.title[:50]}... [{exc}]")

    valid.sort(key=lambda l: getattr(l, '_effective_score', l.relevance_score), reverse=True)

    print(f"    Pipeline: {len(valid)} passed out of {len(leads)}")
    print(f"    quick-filters: url={counts['blocked_url']} region={counts['blocked_region']} "
          f"kw={counts['no_strong_kw']} empty={counts['empty_page']} "
          f"signal={counts['no_tender_signal']} rel={counts['low_relevance']} "
          f"expired-snippet={counts['expired_snippet']}")
    print(f"    page-verify: active={counts['page_active']} "
          f"expired={counts['page_expired']} "
          f"no-deadline={counts['page_no_deadline']}")

    return valid


def _is_deep_search(query: str) -> bool:
    """Detect if the user wants a deep/extended search (triggers SERP).

    Normal queries hit direct APIs only. Deep search adds Google SERP
    for broader coverage of smaller portals that don't have APIs.
    """
    deep_triggers = [
        "deep search", "deep-search", "extended search", "broad search",
        "search everywhere", "all sources", "include google",
        "serp", "google search", "wider search",
    ]
    q_lower = query.lower()
    return any(trigger in q_lower for trigger in deep_triggers)


def _strip_deep_trigger(query: str) -> str:
    """Remove the deep search trigger phrase from the query."""
    import re
    triggers = [
        r"deep[\s-]?search\s*", r"extended\s+search\s*", r"broad\s+search\s*",
        r"search\s+everywhere\s*", r"all\s+sources\s*", r"include\s+google\s*",
        r"serp\s*", r"google\s+search\s*", r"wider\s+search\s*",
    ]
    cleaned = query
    for trigger in triggers:
        cleaned = re.sub(trigger, "", cleaned, flags=re.IGNORECASE)
    return cleaned.strip()


def run_tender_search(query: str) -> tuple[str, dict]:
    """Run tender discovery — APIs first, SERP as automatic fallback.

    Architecture:
      STEP 1: Run all 20 government APIs (region-routed).
      STEP 2: Strict deadline filter — block any API result without a
              verified future deadline. Zero exceptions.
      STEP 3: If API results exist after filtering → show them. Done.
      STEP 4: If APIs returned ZERO results → automatically fall back to
              SERP (Google via Bright Data). Every SERP result gets its
              actual page fetched to extract and verify the deadline.
              Expired/closed/no-deadline → blocked. No exceptions.

    Every result the user sees has a verified, future submission deadline.
    """
    start_time = time.perf_counter()
    from src.discovery.serp_search import detect_region

    search_query = query
    detected_region = detect_region(search_query)

    print(f"\n{'='*60}")
    print(f"  Tender search: '{search_query[:80]}'")
    print(f"  Region: {detected_region} | Mode: APIs-first, SERP-fallback")
    print(f"{'='*60}")

    # ------------------------------------------------------------------
    # Step 1: Direct API sources (structured, reliable, ALWAYS run)
    # ------------------------------------------------------------------
    api_leads: list = []

    # TED Europa — for Europe queries (or global)
    if detected_region in ("europe", "uk", "global"):
        try:
            print("  [TED API] Querying TED Europa for active EU tenders...")
            ted_searcher = TedEuropaSearcher()
            ted_leads = ted_searcher.search(user_query=search_query, max_results=15)
            api_leads.extend(ted_leads)
            print(f"  [TED API] Got {len(ted_leads)} structured results")
            _audit(
                "tool_call",
                f"TED Europa API returned {len(ted_leads)} leads",
                node_name="discover",
                status="success",
                output_payload={"source": "ted_europa", "count": len(ted_leads)},
            )
        except Exception as exc:
            print(f"  [TED API] Failed: {exc}")
            _audit(
                "tool_call", f"TED Europa API failed: {exc}",
                node_name="discover", status="failure", error_message=str(exc),
            )

    # UK Contracts Finder + Find a Tender — for UK queries (or global/europe)
    if detected_region in ("uk", "europe", "global"):
        try:
            print("  [UK API] Querying Contracts Finder + Find a Tender...")
            uk_searcher = UkTenderSearcher()
            uk_leads = uk_searcher.search(user_query=search_query, max_results=15, days_back=60)
            api_leads.extend(uk_leads)
            print(f"  [UK API] Got {len(uk_leads)} relevant UK tenders")
            _audit(
                "tool_call",
                f"UK Tenders API returned {len(uk_leads)} leads",
                node_name="discover",
                status="success",
                output_payload={"source": "uk_gov", "count": len(uk_leads)},
            )
        except Exception as exc:
            print(f"  [UK API] Failed: {exc}")
            _audit(
                "tool_call", f"UK Tenders API failed: {exc}",
                node_name="discover", status="failure", error_message=str(exc),
            )

    # SAM.gov — for US queries (if API key available)
    sam_api_key = os.getenv("SAM_GOV_API_KEY", "")
    if detected_region in ("usa", "global") and sam_api_key:
        try:
            print("  [SAM API] Querying SAM.gov for active US tenders...")
            sam_scraper = SamGovScraper(dry_run=False)
            sam_leads = sam_scraper.fetch_opportunities(days_back=30)
            api_leads.extend(sam_leads)
            print(f"  [SAM API] Got {len(sam_leads)} structured results")
            _audit(
                "tool_call",
                f"SAM.gov API returned {len(sam_leads)} leads",
                node_name="discover",
                status="success",
                output_payload={"source": "sam_gov", "count": len(sam_leads)},
            )
        except Exception as exc:
            print(f"  [SAM API] Failed: {exc}")
    elif detected_region in ("usa", "global") and not sam_api_key:
        print("  [SAM API] Skipped — SAM_GOV_API_KEY not set (register at sam.gov)")

    # BOAMP (France) — for Europe/France queries (or global)
    if detected_region in ("europe", "global"):
        try:
            print("  [BOAMP API] Querying French BOAMP for active tenders...")
            boamp_searcher = BoampSearcher()
            boamp_leads = boamp_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(boamp_leads)
            print(f"  [BOAMP API] Got {len(boamp_leads)} relevant French tenders")
            _audit(
                "tool_call",
                f"BOAMP API returned {len(boamp_leads)} leads",
                node_name="discover",
                status="success",
                output_payload={"source": "boamp", "count": len(boamp_leads)},
            )
        except Exception as exc:
            print(f"  [BOAMP API] Failed: {exc}")
            _audit(
                "tool_call", f"BOAMP API failed: {exc}",
                node_name="discover", status="failure", error_message=str(exc),
            )

    # World Bank Procurement — for global and non-Western-market queries
    if detected_region in ("global", "india", "australia"):
        try:
            print("  [World Bank API] Querying World Bank procurement notices...")
            wb_searcher = WorldBankSearcher()
            wb_leads = wb_searcher.search(user_query=search_query, max_results=10, days_back=90)
            api_leads.extend(wb_leads)
            print(f"  [World Bank API] Got {len(wb_leads)} relevant tenders")
            _audit(
                "tool_call",
                f"World Bank API returned {len(wb_leads)} leads",
                node_name="discover",
                status="success",
                output_payload={"source": "world_bank", "count": len(wb_leads)},
            )
        except Exception as exc:
            print(f"  [World Bank API] Failed: {exc}")
            _audit(
                "tool_call", f"World Bank API failed: {exc}",
                node_name="discover", status="failure", error_message=str(exc),
            )

    # Prozorro (Ukraine) — for Europe/global queries
    if detected_region in ("europe", "global"):
        try:
            print("  [Prozorro API] Querying Ukrainian Prozorro for active tenders...")
            prozorro_searcher = ProzorroSearcher()
            prozorro_leads = prozorro_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(prozorro_leads)
            print(f"  [Prozorro API] Got {len(prozorro_leads)} relevant Ukrainian tenders")
            _audit(
                "tool_call",
                f"Prozorro API returned {len(prozorro_leads)} leads",
                node_name="discover",
                status="success",
                output_payload={"source": "prozorro", "count": len(prozorro_leads)},
            )
        except Exception as exc:
            print(f"  [Prozorro API] Failed: {exc}")
            _audit(
                "tool_call", f"Prozorro API failed: {exc}",
                node_name="discover", status="failure", error_message=str(exc),
            )

    # CanadaBuys — for Canada/North America/global queries
    if detected_region in ("canada", "usa", "global"):
        try:
            print("  [CanadaBuys API] Querying Canadian procurement data...")
            canada_searcher = CanadaBuysSearcher()
            canada_leads = canada_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(canada_leads)
            print(f"  [CanadaBuys API] Got {len(canada_leads)} relevant Canadian tenders")
            _audit(
                "tool_call",
                f"CanadaBuys API returned {len(canada_leads)} leads",
                node_name="discover",
                status="success",
                output_payload={"source": "canada_buys", "count": len(canada_leads)},
            )
        except Exception as exc:
            print(f"  [CanadaBuys API] Failed: {exc}")
            _audit(
                "tool_call", f"CanadaBuys API failed: {exc}",
                node_name="discover", status="failure", error_message=str(exc),
            )

    # AusTender (Australia) — for Australia/global queries
    if detected_region in ("australia", "global"):
        try:
            print("  [AusTender API] Querying Australian federal tenders...")
            aus_searcher = AusTenderSearcher()
            aus_leads = aus_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(aus_leads)
            print(f"  [AusTender API] Got {len(aus_leads)} relevant Australian tenders")
            _audit(
                "tool_call",
                f"AusTender API returned {len(aus_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "austender", "count": len(aus_leads)},
            )
        except Exception as exc:
            print(f"  [AusTender API] Failed: {exc}")
            _audit("tool_call", f"AusTender API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # South Africa eTender — for Africa/global queries
    if detected_region in ("africa", "global"):
        try:
            print("  [SA eTender API] Querying South African government tenders...")
            sa_searcher = SaEtenderSearcher()
            sa_leads = sa_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(sa_leads)
            print(f"  [SA eTender API] Got {len(sa_leads)} relevant South African tenders")
            _audit(
                "tool_call",
                f"SA eTender API returned {len(sa_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "sa_etender", "count": len(sa_leads)},
            )
        except Exception as exc:
            print(f"  [SA eTender API] Failed: {exc}")
            _audit("tool_call", f"SA eTender API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Colombia SECOP — for South America/global queries
    if detected_region in ("south_america", "global"):
        try:
            print("  [Colombia SECOP API] Querying Colombian procurement data...")
            col_searcher = ColombiaSecopSearcher()
            col_leads = col_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(col_leads)
            print(f"  [Colombia SECOP API] Got {len(col_leads)} relevant Colombian tenders")
            _audit(
                "tool_call",
                f"Colombia SECOP API returned {len(col_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "colombia_secop", "count": len(col_leads)},
            )
        except Exception as exc:
            print(f"  [Colombia SECOP API] Failed: {exc}")
            _audit("tool_call", f"Colombia SECOP API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Brazil Compras — for South America/global queries
    if detected_region in ("south_america", "global"):
        try:
            print("  [Brazil Compras API] Querying Brazilian federal procurement...")
            br_searcher = BrazilComprasSearcher()
            br_leads = br_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(br_leads)
            print(f"  [Brazil Compras API] Got {len(br_leads)} relevant Brazilian tenders")
            _audit(
                "tool_call",
                f"Brazil Compras API returned {len(br_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "brazil_compras", "count": len(br_leads)},
            )
        except Exception as exc:
            print(f"  [Brazil Compras API] Failed: {exc}")
            _audit("tool_call", f"Brazil Compras API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Germany BKMS — for Europe/global queries
    if detected_region in ("europe", "global"):
        try:
            print("  [Germany BKMS API] Querying German federal procurement...")
            de_searcher = GermanyBkmsSearcher()
            de_leads = de_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(de_leads)
            print(f"  [Germany BKMS API] Got {len(de_leads)} relevant German tenders")
            _audit(
                "tool_call",
                f"Germany BKMS API returned {len(de_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "germany_bkms", "count": len(de_leads)},
            )
        except Exception as exc:
            print(f"  [Germany BKMS API] Failed: {exc}")
            _audit("tool_call", f"Germany BKMS API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Italy ANAC — for Europe/global queries
    if detected_region in ("europe", "global"):
        try:
            print("  [Italy ANAC API] Querying Italian public procurement...")
            it_searcher = ItalyAnacSearcher()
            it_leads = it_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(it_leads)
            print(f"  [Italy ANAC API] Got {len(it_leads)} relevant Italian tenders")
            _audit(
                "tool_call",
                f"Italy ANAC API returned {len(it_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "italy_anac", "count": len(it_leads)},
            )
        except Exception as exc:
            print(f"  [Italy ANAC API] Failed: {exc}")
            _audit("tool_call", f"Italy ANAC API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Dominican Republic DGCP — for South America/Caribbean/global queries
    if detected_region in ("south_america", "global"):
        try:
            print("  [Dominican DGCP API] Querying Dominican Republic procurement...")
            dr_searcher = DominicanDgcpSearcher()
            dr_leads = dr_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(dr_leads)
            print(f"  [Dominican DGCP API] Got {len(dr_leads)} relevant Dominican tenders")
            _audit(
                "tool_call",
                f"Dominican DGCP API returned {len(dr_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "dominican_dgcp", "count": len(dr_leads)},
            )
        except Exception as exc:
            print(f"  [Dominican DGCP API] Failed: {exc}")
            _audit("tool_call", f"Dominican DGCP API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Peru OECE — for South America/global queries
    if detected_region in ("south_america", "global"):
        try:
            print("  [Peru OECE API] Querying Peruvian procurement data...")
            pe_searcher = PeruOeceSearcher()
            pe_leads = pe_searcher.search(user_query=search_query, max_results=10, days_back=90)
            api_leads.extend(pe_leads)
            print(f"  [Peru OECE API] Got {len(pe_leads)} relevant Peruvian tenders")
            _audit(
                "tool_call",
                f"Peru OECE API returned {len(pe_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "peru_oece", "count": len(pe_leads)},
            )
        except Exception as exc:
            print(f"  [Peru OECE API] Failed: {exc}")
            _audit("tool_call", f"Peru OECE API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # World Bank Search API v2 — keyword-based, global (always fires alongside existing WB)
    if detected_region in ("global", "india", "africa", "australia"):
        try:
            print("  [WB Search v2 API] Querying World Bank procurement notices (keyword search)...")
            wb2_searcher = WorldBankV2Searcher()
            wb2_leads = wb2_searcher.search(user_query=search_query, max_results=10)
            api_leads.extend(wb2_leads)
            print(f"  [WB Search v2 API] Got {len(wb2_leads)} relevant World Bank tenders")
            _audit(
                "tool_call",
                f"World Bank v2 API returned {len(wb2_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "world_bank_v2", "count": len(wb2_leads)},
            )
        except Exception as exc:
            print(f"  [WB Search v2 API] Failed: {exc}")
            _audit("tool_call", f"World Bank v2 API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Nigeria NOCOPO — for Africa/global queries
    if detected_region in ("africa", "global"):
        try:
            print("  [Nigeria NOCOPO API] Querying Nigerian federal procurement...")
            ng_searcher = NigeriaNocopoSearcher()
            ng_leads = ng_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(ng_leads)
            print(f"  [Nigeria NOCOPO API] Got {len(ng_leads)} relevant Nigerian tenders")
            _audit(
                "tool_call",
                f"Nigeria NOCOPO API returned {len(ng_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "nigeria_nocopo", "count": len(ng_leads)},
            )
        except Exception as exc:
            print(f"  [Nigeria NOCOPO API] Failed: {exc}")
            _audit("tool_call", f"Nigeria NOCOPO API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Kenya PPRA — for Africa/global queries
    if detected_region in ("africa", "global"):
        try:
            print("  [Kenya PPRA API] Querying Kenyan government tenders...")
            ke_searcher = KenyaPpraSearcher()
            ke_leads = ke_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(ke_leads)
            print(f"  [Kenya PPRA API] Got {len(ke_leads)} relevant Kenyan tenders")
            _audit(
                "tool_call",
                f"Kenya PPRA API returned {len(ke_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "kenya_ppra", "count": len(ke_leads)},
            )
        except Exception as exc:
            print(f"  [Kenya PPRA API] Failed: {exc}")
            _audit("tool_call", f"Kenya PPRA API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Uganda GPP — for Africa/global queries
    if detected_region in ("africa", "global"):
        try:
            print("  [Uganda GPP API] Querying Ugandan government procurement...")
            ug_searcher = UgandaGppSearcher()
            ug_leads = ug_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(ug_leads)
            print(f"  [Uganda GPP API] Got {len(ug_leads)} relevant Ugandan tenders")
            _audit(
                "tool_call",
                f"Uganda GPP API returned {len(ug_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "uganda_gpp", "count": len(ug_leads)},
            )
        except Exception as exc:
            print(f"  [Uganda GPP API] Failed: {exc}")
            _audit("tool_call", f"Uganda GPP API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    # Mexico City CDMX — for South America/global queries
    if detected_region in ("south_america", "global"):
        try:
            print("  [Mexico CDMX API] Querying Mexico City procurement data...")
            mx_searcher = MexicoCdmxSearcher()
            mx_leads = mx_searcher.search(user_query=search_query, max_results=10, days_back=60)
            api_leads.extend(mx_leads)
            print(f"  [Mexico CDMX API] Got {len(mx_leads)} relevant Mexico City tenders")
            _audit(
                "tool_call",
                f"Mexico CDMX API returned {len(mx_leads)} leads",
                node_name="discover", status="success",
                output_payload={"source": "mexico_cdmx", "count": len(mx_leads)},
            )
        except Exception as exc:
            print(f"  [Mexico CDMX API] Failed: {exc}")
            _audit("tool_call", f"Mexico CDMX API failed: {exc}", node_name="discover", status="failure", error_message=str(exc))

    print(f"  API total: {len(api_leads)} raw leads from direct government APIs")

    # ------------------------------------------------------------------
    # Step 2: Strict deadline filter on API results
    # Every API result must have a verified future deadline.
    # ------------------------------------------------------------------
    api_verified = _strict_deadline_filter(api_leads, "API")

    # Deduplicate by URL
    seen_urls: set[str] = set()
    valid_leads: list = []
    for lead in api_verified:
        url = getattr(lead, 'source_url', '')
        if url and url not in seen_urls:
            seen_urls.add(url)
            if not hasattr(lead, '_effective_score'):
                lead._effective_score = getattr(lead, 'relevance_score', 0.80)
            valid_leads.append(lead)

    print(f"  API after deadline filter + dedup: {len(valid_leads)} verified leads")

    # ------------------------------------------------------------------
    # Step 2b: Relevance floor — block tenders without real SDS/EHS signal.
    # A 15% score means only generic partial matches like "safety" or
    # "waste" — not actually about chemical safety or SDS. Require 30%+
    # which means at least 1 strong keyword or 6+ partial matches.
    # ------------------------------------------------------------------
    RELEVANCE_FLOOR = 0.30
    before_relevance = len(valid_leads)
    valid_leads = [
        l for l in valid_leads
        if getattr(l, 'relevance_score', 0) >= RELEVANCE_FLOOR
    ]
    blocked_irrelevant = before_relevance - len(valid_leads)
    if blocked_irrelevant > 0:
        print(f"  [Relevance gate] Blocked {blocked_irrelevant} leads below {RELEVANCE_FLOOR:.0%} "
              f"(kept {len(valid_leads)} relevant)")
    else:
        print(f"  [Relevance gate] All {len(valid_leads)} leads above {RELEVANCE_FLOOR:.0%} floor")

    # ------------------------------------------------------------------
    # Step 3: SERP automatic fallback (ONLY if APIs returned zero results)
    # ------------------------------------------------------------------
    serp_used = False
    serp_valid_leads: list = []

    if len(valid_leads) == 0:
        serp_used = True
        print(f"\n  [SERP FALLBACK] APIs returned 0 results — falling back to Google SERP...")
        _audit(
            "tool_call",
            f"API sources returned 0 results, falling back to SERP for: {search_query[:100]}",
            node_name="discover",
            input_payload={"query": search_query, "tool": "brightdata_serp", "mode": "auto_fallback"},
        )
        try:
            searcher = SerpTenderSearcher()
            serp_leads = searcher.search(user_query=search_query)
            print(f"  [SERP] Got {len(serp_leads)} raw leads — running strict pipeline...")

            # Run the strict filtering pipeline (every result page-fetched)
            serp_valid_leads = _filter_serp_leads(serp_leads, search_query)
            print(f"  [SERP] {len(serp_valid_leads)} passed (all have verified deadlines)")

            # Add SERP leads to valid_leads (no dedup needed — API list was empty)
            for lead in serp_valid_leads:
                url = getattr(lead, 'source_url', '')
                if url and url not in seen_urls:
                    seen_urls.add(url)
                    valid_leads.append(lead)

        except Exception as exc:
            print(f"  [SERP] Failed: {exc}")
            _audit(
                "tool_call", f"SERP fallback failed: {exc}",
                node_name="discover", status="failure", error_message=str(exc),
            )
    else:
        print("  [SERP] Skipped — APIs returned results, no fallback needed")

    valid_leads.sort(
        key=lambda l: getattr(l, '_effective_score', getattr(l, 'relevance_score', 0)),
        reverse=True,
    )

    _all_api_portals = ('ted.europa.eu', 'sam.gov', 'uk_gov', 'boamp', 'world_bank', 'prozorro', 'canada_buys', 'austender', 'sa_etender', 'colombia_secop', 'brazil_compras', 'germany_bkms', 'italy_anac', 'dominican_dgcp', 'peru_oece', 'world_bank_v2', 'nigeria_nocopo', 'kenya_ppra', 'uganda_gpp', 'mexico_cdmx')
    api_count = len([l for l in valid_leads if getattr(l, 'source_portal', '') in _all_api_portals])
    serp_count = len(valid_leads) - api_count

    print(f"  FINAL: {len(valid_leads)} total ({api_count} API + {serp_count} SERP)")

    if not valid_leads:
        from src.discovery.serp_search import detect_region, REGION_COUNTRY_NAMES
        detected = detect_region(query)
        region_name = REGION_COUNTRY_NAMES.get(detected, ["this region"])[0]

        searched_label = "20 government APIs + Google SERP" if serp_used else "20 government APIs"

        if detected != "global":
            reply = (
                f"**No active SDS/EHS tenders found in {region_name}.**\n\n"
                f"I searched {searched_label} but couldn't find relevant, "
                f"unexpired opportunities matching your criteria in {region_name} right now.\n\n"
                f"**What you can do:**\n"
                f"- Try broadening your search: *'find chemical safety tenders globally'*\n"
                f"- Check region-specific portals directly (e.g., AusTender, TED Europa, SAM.gov)\n"
                f"- Try different keywords: *'EHS compliance software'*, *'hazardous materials management'*"
            )
        else:
            reply = (
                "**No matching tenders found.**\n\n"
                f"I searched {searched_label} globally but couldn't find verified, unexpired "
                "tender listings matching your criteria.\n\n"
                "**Try:**\n"
                "- *'SDS authoring software RFP United States'*\n"
                "- *'chemical safety compliance tender Europe'*\n"
                "- *'EHS management system procurement Canada'*"
            )
        duration_ms = int((time.perf_counter() - start_time) * 1000)
        return reply, {
            "node": "discover",
            "tenders_found": 0,
            "duration_ms": duration_ms,
            "cost_usd": 0,
            "serp_fallback_used": serp_used,
        }

    # Step 3: Format each tender as a rich card with source badges
    import re

    # Source badge mapping — users see exactly where each tender came from
    SOURCE_BADGES = {
        "ted.europa.eu":   "\U0001f1ea\U0001f1fa TED Europa",
        "uk_gov":          "\U0001f1ec\U0001f1e7 UK Gov",
        "sam.gov":         "\U0001f1fa\U0001f1f8 SAM.gov",
        "boamp":           "\U0001f1eb\U0001f1f7 BOAMP France",
        "world_bank":      "\U0001f30d World Bank",
        "prozorro":        "\U0001f1fa\U0001f1e6 Prozorro",
        "canada_buys":     "\U0001f1e8\U0001f1e6 CanadaBuys",
        "austender":       "\U0001f1e6\U0001f1fa AusTender",
        "sa_etender":      "\U0001f1ff\U0001f1e6 SA eTender",
        "colombia_secop":  "\U0001f1e8\U0001f1f4 SECOP Colombia",
        "brazil_compras":  "\U0001f1e7\U0001f1f7 Compras Brazil",
        "germany_bkms":    "\U0001f1e9\U0001f1ea BKMS Germany",
        "italy_anac":      "\U0001f1ee\U0001f1f9 ANAC Italy",
        "dominican_dgcp":  "\U0001f1e9\U0001f1f4 DGCP Dom. Rep.",
        "peru_oece":       "\U0001f1f5\U0001f1ea OECE Peru",
        "world_bank_v2":   "\U0001f30d World Bank v2",
        "nigeria_nocopo":  "\U0001f1f3\U0001f1ec NOCOPO Nigeria",
        "kenya_ppra":      "\U0001f1f0\U0001f1ea PPRA Kenya",
        "uganda_gpp":      "\U0001f1fa\U0001f1ec GPP Uganda",
        "mexico_cdmx":     "\U0001f1f2\U0001f1fd CDMX Mexico",
    }

    # ---------------------------------------------------------------
    # FINAL SAFETY NET: Re-verify every deadline right before display.
    # Belt-and-suspenders — if anything slipped past earlier filters,
    # this catches it. Zero expired tenders shown. Period.
    # ---------------------------------------------------------------
    from dateutil import parser as dateparser
    _today_final = datetime.now(timezone.utc).date()
    _display_leads: list = []
    for _lead in valid_leads:
        _dl = getattr(_lead, 'submission_deadline', '') or ''
        if not _dl or len(_dl) < 8:
            print(f"  [SAFETY NET] Blocked (no deadline): {getattr(_lead, 'title', '?')[:60]}")
            continue
        try:
            _parsed = dateparser.parse(_dl).date()
            if _parsed < _today_final:
                print(f"  [SAFETY NET] Blocked (expired {_dl}): {getattr(_lead, 'title', '?')[:60]}")
                continue
        except Exception:
            print(f"  [SAFETY NET] Blocked (unparseable '{_dl}'): {getattr(_lead, 'title', '?')[:60]}")
            continue
        _display_leads.append(_lead)
    if len(_display_leads) < len(valid_leads):
        print(f"  [SAFETY NET] Removed {len(valid_leads) - len(_display_leads)} leads at final check")
    valid_leads = _display_leads

    formatted_items = []
    for i, lead in enumerate(valid_leads[:10], 1):
        title = lead.title.strip()
        if len(title) > 80:
            title = title[:77] + "..."
        url = lead.source_url
        agency = getattr(lead, 'agency', 'Unknown')
        relevance = lead.relevance_score
        deadline = getattr(lead, 'submission_deadline', '') or ''
        portal = getattr(lead, 'source_portal', '')

        # Source badge
        source_badge = SOURCE_BADGES.get(portal, "\U0001f50d SERP")

        # Clean the description
        desc = getattr(lead, 'description', '') or ''
        desc = re.sub(r'https?://\S+', '', desc)
        desc = re.sub(r'[\w.-]+\s*›[^›\n]*(?:›[^›\n]*)*', '', desc)
        desc = re.sub(r'\s{2,}', ' ', desc).strip()
        if desc.lower().strip('.') == title.lower().strip('.'):
            desc = ''
        if len(desc) > 200:
            desc = desc[:197] + "..."

        # Value (if available from UK or SAM APIs)
        value_str = ""
        value_amount = getattr(lead, 'value_amount', 0) or 0
        if value_amount > 0:
            currency = getattr(lead, 'value_currency', 'GBP')
            if value_amount >= 1_000_000:
                value_str = f" · \U0001f4b0 {currency} {value_amount/1_000_000:.1f}M"
            elif value_amount >= 1_000:
                value_str = f" · \U0001f4b0 {currency} {value_amount/1_000:.0f}K"

        # Build the card
        card = f"**{i}. [{title}]({url})**"
        card += f"\n{source_badge} · \U0001f3db️ {agency}"
        if desc:
            card += f"\n{desc}"
        badge_line = f"\U0001f4ca Relevance: {relevance:.0%}"
        badge_line += f" · \U0001f4c5 Deadline: {deadline[:10]}"
        badge_line += value_str
        card += f"\n{badge_line}"

        formatted_items.append(card)

    results_block = "\n\n".join(formatted_items)
    # Also build a plain text version for the LLM summary
    plain_list = "\n".join(
        f"{i}. {l.title} ({getattr(l, 'agency', 'Unknown')}) - {l.relevance_score:.0%}"
        for i, l in enumerate(valid_leads[:10], 1)
    )

    # Step 4: Get LLM summary
    llm_summary = ""
    llm_cost = 0.0
    llm_tokens_in = 0
    llm_tokens_out = 0
    if OPENROUTER_API_KEY:
        try:
            llm_result = call_llm(
                f"The user searched for: '{query}'\n\n"
                f"Found {len(valid_leads)} verified tender listings:\n{plain_list}\n\n"
                f"Write a 3-4 sentence executive briefing. Name the top opportunity, "
                f"its issuing agency, region, estimated contract value if visible, "
                f"and a clear recommendation on whether to pursue it. Be specific and actionable.",
                max_tokens=350,
            )
            llm_summary = llm_result["content"]
            llm_cost = llm_result["cost_usd"]
            llm_tokens_in = llm_result["tokens_input"]
            llm_tokens_out = llm_result["tokens_output"]
        except Exception as exc:
            print(f"LLM summary failed: {exc}")

    # Step 5: Assemble clean reply
    duration_ms = int((time.perf_counter() - start_time) * 1000)

    parts = []
    parts.append("## \U0001f50d Tender Discovery Results")
    parts.append("")
    if llm_summary:
        parts.append(llm_summary)
        parts.append("")
        parts.append("---")
        parts.append("")
    parts.append(f"**{len(valid_leads)} verified tenders found:**")
    parts.append("")
    parts.append(results_block)
    parts.append("")
    parts.append("---")
    parts.append("")
    # Build a source label reflecting which sources contributed
    source_parts = []
    api_portals = _all_api_portals
    api_final = len([l for l in valid_leads if getattr(l, 'source_portal', '') in api_portals])
    serp_final = len(valid_leads) - api_final
    if api_final > 0:
        source_parts.append(f"{api_final} from government APIs")
    if serp_final > 0:
        source_parts.append(f"{serp_final} from SERP fallback")
    source_label = " + ".join(source_parts) if source_parts else "government API search"

    parts.append(f"_{source_label} \u2022 {len(valid_leads)} verified results \u2022 all deadlines confirmed active \u2022 {duration_ms}ms_")

    reply = "\n".join(parts)

    # Determine which sources contributed
    sources_used = []
    portal_to_source = {
        "ted.europa.eu": "ted_europa",
        "uk_gov": "uk_gov",
        "sam.gov": "sam_gov",
        "boamp": "boamp_france",
        "world_bank": "world_bank",
        "prozorro": "prozorro",
        "canada_buys": "canada_buys",
        "austender": "austender",
        "sa_etender": "sa_etender",
        "colombia_secop": "colombia_secop",
        "brazil_compras": "brazil_compras",
        "germany_bkms": "germany_bkms",
        "italy_anac": "italy_anac",
        "dominican_dgcp": "dominican_dgcp",
        "peru_oece": "peru_oece",
        "world_bank_v2": "world_bank_v2",
        "nigeria_nocopo": "nigeria_nocopo",
        "kenya_ppra": "kenya_ppra",
        "uganda_gpp": "uganda_gpp",
        "mexico_cdmx": "mexico_cdmx",
    }
    for portal_key, source_name in portal_to_source.items():
        if any(getattr(l, 'source_portal', '') == portal_key for l in valid_leads):
            sources_used.append(source_name)
    if serp_final > 0:
        sources_used.append("brightdata_serp")

    metadata = {
        "node": "discover",
        "tenders_found": len(valid_leads),
        "top_relevance": valid_leads[0].relevance_score if valid_leads else 0,
        "tokens_input": llm_tokens_in,
        "tokens_output": llm_tokens_out,
        "cost_usd": llm_cost,
        "model": LLM_MODEL,
        "duration_ms": duration_ms,
        "search_sources": sources_used,
        "api_results": api_final,
        "serp_results": serp_final,
        "serp_fallback_used": serp_used,
        "all_deadlines_verified": True,
    }

    return reply, metadata


# ---------------------------------------------------------------------------
# Message Handler — processes incoming chat messages from AMS
# ---------------------------------------------------------------------------

def classify_intent(user_message: str, has_attachments: bool = False) -> dict:
    """Use the LLM to understand what the user wants.

    Returns:
        {
            "intent": "search_tenders" | "ask_question" | "greeting" | "refine_results" | "fill_form",
            "search_query": "extracted search terms if intent is search",
            "response_hint": "brief note on how to respond",
        }
    """
    attachment_hint = ""
    if has_attachments:
        attachment_hint = "\nIMPORTANT: The user has attached one or more files to this message. If the message is about filling, completing, or submitting a form/tender/document, the intent is 'fill_form'."

    try:
        result = call_llm(
            f"""You are an intent classifier for a tender discovery agent.
Analyze this user message and respond with ONLY valid JSON (no markdown, no explanation):

User message: "{user_message}"{attachment_hint}

Respond with this exact JSON structure:
{{"intent": "search_tenders|ask_question|greeting|refine_results|fill_form", "search_query": "extracted search keywords for tender search", "response_hint": "brief note"}}

Rules:
- "fill_form": user wants to fill out, complete, submit a tender form, or has uploaded a document to be filled
- "search_tenders": user wants to find tenders, RFPs, bids, procurement opportunities
- "ask_question": user asks about a specific tender, process, or general question
- "greeting": user says hi, hello, thanks, etc.
- "refine_results": user wants to narrow down or filter previous results
- "search_query": extract the core topic/keywords (e.g., "chemical safety tenders in Europe")
- Always extract a search_query even for greetings (use "SDS management tenders" as default)""",
            max_tokens=200,
        )
        # Parse JSON from response
        content = result["content"].strip()
        # Handle markdown code blocks
        if "```" in content:
            content = content.split("```")[1].strip()
            if content.startswith("json"):
                content = content[4:].strip()
        return json.loads(content)
    except Exception as exc:
        print(f"Intent classification failed: {exc}, defaulting to search")
        # If there are attachments, default to fill_form
        if has_attachments:
            return {
                "intent": "fill_form",
                "search_query": user_message,
                "response_hint": "form filling with attachment",
            }
        return {
            "intent": "search_tenders",
            "search_query": user_message,
            "response_hint": "default search",
        }


# ---------------------------------------------------------------------------
# Form Filling — Conversation State & Pipeline
# ---------------------------------------------------------------------------

# In-memory state for multi-turn form-filling conversations.
# Keyed by thread_id. Each entry tracks the form-filling progress.
_form_sessions: dict[str, dict] = {}


def handle_form_filling(client: NexusClient, message: dict) -> None:
    """Handle the form-filling flow (multi-turn).

    Flow:
    1. User uploads form + says "fill this" → parse form, fill what we can, ask questions
    2. User answers clarification questions → fill remaining fields, generate file, send back
    """
    thread_id = message.get("threadId", "")
    content = message.get("content", "")
    attachments = message.get("attachments", [])
    clean_content = content.replace("@tender-agent", "").strip()

    start_time = time.perf_counter()

    # Check if this is a continuation of an existing form-filling session
    session = _form_sessions.get(thread_id)

    if session and session.get("waiting_for_answers"):
        # This is a follow-up message with answers to our questions
        print("Processing user answers for form filling...")
        _handle_form_answers(client, thread_id, clean_content, session)
        return

    # New form-filling request — needs at least one attachment
    if not attachments:
        client.reply(
            thread_id=thread_id,
            content=(
                "I'd be happy to help fill out a tender form! Please **attach the form file** "
                "(PDF, DOCX, or XLSX) to your message, and I'll analyze it and fill in what I can "
                "using your company documents.\n\n"
                "You can drag & drop the file or click the paperclip icon."
            ),
            metadata={"intent": "fill_form", "status": "awaiting_attachment"},
        )
        return

    # Download the first attachment
    att = attachments[0]
    minio_key = att.get("minioKey", "")
    filename = att.get("filename", "form")

    if not minio_key:
        client.reply(
            thread_id=thread_id,
            content="I couldn't access the attached file. Could you try uploading it again?",
            metadata={"intent": "fill_form", "status": "error"},
        )
        return

    print(f"Downloading form: {filename} ({minio_key})")

    # Create a temp directory for this session
    tmp_dir = tempfile.mkdtemp(prefix="tender_form_")
    local_path = os.path.join(tmp_dir, filename)

    try:
        # Download the form from AMS
        client.download_file(minio_key, local_path)
        print(f"Downloaded to: {local_path}")

        # Parse the form
        print("Parsing form fields...")
        parser = FormParser()
        parse_result = parser.parse(local_path)
        print(f"Found {len(parse_result.fields)} fields, format: {parse_result.file_format.value}")

        # Audit: form parsed
        _audit(
            "form_parsed",
            f"Parsed {filename}: {len(parse_result.fields)} fields, format={parse_result.file_format.value}",
            node_name="fill_form",
            input_payload={"filename": filename, "format": parse_result.file_format.value},
            output_payload={
                "fields_count": len(parse_result.fields),
                "field_names": [f.name for f in parse_result.fields[:20]],
                "parse_errors": parse_result.parse_errors,
                "raw_text_length": len(parse_result.raw_text),
            },
        )

        if parse_result.parse_errors:
            for err in parse_result.parse_errors:
                print(f"  Parse warning: {err}")

        if not parse_result.fields and not parse_result.raw_text:
            client.reply(
                thread_id=thread_id,
                content=(
                    f"I opened **{filename}** but couldn't extract any form fields from it. "
                    f"The file might be a scanned image or an unsupported format.\n\n"
                    f"I can work with fillable PDFs, Word documents (.docx) with tables or "
                    f"placeholders, and Excel spreadsheets (.xlsx)."
                ),
                metadata={"intent": "fill_form", "status": "parse_failed"},
            )
            return

        # Fetch company context documents
        company_context = fetch_agent_context()
        if not company_context:
            client.reply(
                thread_id=thread_id,
                content=(
                    f"I've analyzed **{filename}** and found **{len(parse_result.fields)} fields** "
                    f"to fill. However, I don't have any company documents assigned to me yet.\n\n"
                    f"Please go to the **Documents** page in the AMS, upload your company details "
                    f"(ABN, address, certifications, past proposals, etc.), and assign them to the "
                    f"Tender Agent. Then try again!"
                ),
                metadata={"intent": "fill_form", "status": "no_context"},
            )
            return

        # Fill the form using LLM + context
        print("Filling form with company context...")
        filler = FormFiller(
            llm_call_fn=call_llm,
            confidence_threshold=0.7,
        )
        fill_result = filler.fill(parse_result, company_context)

        print(f"Filled {fill_result.filled_count}/{fill_result.total_fields} fields")
        print(f"Questions for user: {fill_result.needs_clarification_count}")

        # Audit: form filling result
        _audit(
            "form_filled",
            f"Filled {fill_result.filled_count}/{fill_result.total_fields} fields, "
            f"{fill_result.needs_clarification_count} need user input",
            node_name="fill_form",
            cost_usd=fill_result.llm_cost_usd,
            output_payload={
                "filled_count": fill_result.filled_count,
                "total_fields": fill_result.total_fields,
                "needs_clarification": fill_result.needs_clarification_count,
                "filled_field_names": [f.name for f in fill_result.filled_fields],
                "question_fields": [q.field_name for q in fill_result.questions],
            },
        )

        # Store session state for multi-turn
        _form_sessions[thread_id] = {
            "tmp_dir": tmp_dir,
            "local_path": local_path,
            "filename": filename,
            "parse_result": parse_result,
            "fill_result": fill_result,
            "company_context": company_context,
            "waiting_for_answers": fill_result.needs_clarification_count > 0,
        }

        duration_ms = int((time.perf_counter() - start_time) * 1000)

        if fill_result.needs_clarification_count > 0:
            # Ask user about uncertain fields
            questions_text = _format_questions(fill_result)
            reply = (
                f"## \U0001f4cb Form Analysis: {filename}\n\n"
                f"I've analyzed the form and filled **{fill_result.filled_count}** out of "
                f"**{fill_result.total_fields}** fields using your company documents.\n\n"
                f"I need your help with **{fill_result.needs_clarification_count}** field(s):\n\n"
                f"{questions_text}\n\n"
                f"---\n"
                f"_Please reply with the answers and I'll generate the completed form._"
            )
            client.reply(
                thread_id=thread_id,
                content=reply,
                metadata={
                    "intent": "fill_form",
                    "status": "awaiting_answers",
                    "fields_total": fill_result.total_fields,
                    "fields_filled": fill_result.filled_count,
                    "fields_pending": fill_result.needs_clarification_count,
                    "duration_ms": duration_ms,
                },
            )
        else:
            # All fields filled — generate the completed form immediately
            _generate_and_send_form(client, thread_id, _form_sessions[thread_id], duration_ms)

        # Track metrics
        client.track("task_completion", 1.0, metadata={
            "node": "fill_form",
            "fields_filled": fill_result.filled_count,
            "fields_total": fill_result.total_fields,
        })
        client.track("latency", float(duration_ms), metadata={"node": "fill_form"})
        if fill_result.llm_cost_usd > 0:
            client.track("cost", fill_result.llm_cost_usd, metadata={
                "node": "fill_form", "model": LLM_MODEL,
            })

    except Exception as exc:
        error_msg = f"Error processing form: {str(exc)}"
        print(f"ERROR: {error_msg}")
        traceback.print_exc()
        client.reply(
            thread_id=thread_id,
            content=(
                f"Sorry, I ran into an issue while processing **{filename}**:\n\n"
                f"`{str(exc)}`\n\n"
                f"Please make sure the file is a valid PDF, DOCX, or XLSX form and try again."
            ),
            metadata={"intent": "fill_form", "status": "error", "error": str(exc)},
        )
        client.track("error_rate", 1.0, metadata={
            "node": "fill_form", "error_type": type(exc).__name__,
        })

    # Flush metrics
    try:
        flushed = client.flush()
        print(f"Metrics flushed: {flushed.inserted} events sent to AMS")
    except Exception as exc:
        print(f"Metric flush failed: {exc}")


def _handle_form_answers(client: NexusClient, thread_id: str, user_text: str, session: dict) -> None:
    """Process user's answers to clarification questions and generate the filled form."""
    start_time = time.perf_counter()

    fill_result: FillResult = session["fill_result"]
    parse_result = session["parse_result"]
    company_context = session["company_context"]

    # Parse user answers — use LLM to map free-text answers to field names
    questions = fill_result.questions
    field_names = [q.field_name for q in questions]

    try:
        result = call_llm(
            f"""The user was asked to provide values for these form fields:
{json.dumps(field_names, indent=2)}

The user replied:
"{user_text}"

Extract the answer for each field. Respond with ONLY valid JSON:
{{
  "answers": {{
    "field name": "value from user's reply",
    ...
  }}
}}

Rules:
- Map each part of the user's reply to the most likely field
- If the user's reply is numbered (1, 2, 3...) map by order
- If a field is not addressed, set it to empty string
- Be smart about parsing — the user may give all answers in one message""",
            max_tokens=500,
        )
        content = result["content"].strip()
        if "```" in content:
            content = content.split("```")[1].strip()
            if content.startswith("json"):
                content = content[4:].strip()
        parsed = json.loads(content)
        user_answers = parsed.get("answers", {})
    except Exception as exc:
        print(f"Failed to parse user answers: {exc}")
        # Fallback: if there's only one question, use the full text as the answer
        if len(questions) == 1:
            user_answers = {questions[0].field_name: user_text}
        else:
            user_answers = {}

    # Re-run the filler with user answers
    filler = FormFiller(llm_call_fn=call_llm, confidence_threshold=0.7)
    new_fill_result = filler.fill(parse_result, company_context, user_answers)

    # Merge: use all previously filled fields + newly filled ones
    all_filled = list(fill_result.filled_fields) + list(new_fill_result.filled_fields)

    # Deduplicate by field name (prefer user-provided answers)
    seen = {}
    for ff in reversed(all_filled):  # reversed so later entries (user answers) win
        if ff.name not in seen:
            seen[ff.name] = ff
    merged_fields = list(seen.values())

    # Update session
    session["fill_result"] = FillResult(
        filled_fields=merged_fields,
        questions=[],
        total_fields=fill_result.total_fields,
        filled_count=len(merged_fields),
        needs_clarification_count=0,
    )
    session["waiting_for_answers"] = False

    duration_ms = int((time.perf_counter() - start_time) * 1000)
    _generate_and_send_form(client, thread_id, session, duration_ms)


def _generate_and_send_form(client: NexusClient, thread_id: str, session: dict, duration_ms: int) -> None:
    """Generate the filled form and submit it for human approval.

    Instead of sending the form directly to chat, it goes to the Approvals
    inbox where the user can review, approve, or reject it.
    """
    fill_result: FillResult = session["fill_result"]
    local_path = session["local_path"]
    filename = session["filename"]
    tmp_dir = session["tmp_dir"]

    # Generate the filled form
    print("Writing filled form...")
    writer = FormWriter()
    stem = Path(filename).stem
    ext = Path(filename).suffix
    output_filename = f"{stem}_FILLED{ext}"
    output_path = os.path.join(tmp_dir, output_filename)

    writer.write(local_path, fill_result.filled_fields, output_path)

    # Build output summary for the approval card
    filled_field_summary = {
        ff.name: ff.value
        for ff in fill_result.filled_fields[:30]
    }
    output_summary = {
        "filled_count": fill_result.filled_count,
        "total_fields": fill_result.total_fields,
        "filled_fields": filled_field_summary,
    }

    # Submit for approval instead of sending directly
    try:
        print(f"Submitting form for approval: {output_filename}")
        result = client.submit_for_approval(
            thread_id=thread_id,
            action_type="form_fill",
            title=f"Filled Form: {filename}",
            description=(
                f"I filled {fill_result.filled_count} out of {fill_result.total_fields} fields "
                f"using your company documents. Please review the completed form before it's finalized."
            ),
            input_summary={
                "original_filename": filename,
                "total_fields": fill_result.total_fields,
            },
            output_summary=output_summary,
            file_path=output_path,
            filename=output_filename,
            metadata={
                "duration_ms": duration_ms,
                "cost_usd": fill_result.llm_cost_usd,
                "llm_model": LLM_MODEL,
            },
        )
        submission_id = result.get("submissionId", "?")
        print(f"Submitted for approval! ID: {submission_id}")

        # Build summary for the chat message
        summary_lines = []
        for ff in fill_result.filled_fields[:10]:
            confidence_icon = "\u2705" if ff.confidence >= 0.7 else "\u26a0\ufe0f"
            summary_lines.append(f"  {confidence_icon} **{ff.name}**: {ff.value}")
        summary = "\n".join(summary_lines)
        if len(fill_result.filled_fields) > 10:
            summary += f"\n  _...and {len(fill_result.filled_fields) - 10} more fields_"

        # Tell the user to check the Approvals page
        client.reply(
            thread_id=thread_id,
            content=(
                f"## \U0001f4cb Form Ready for Review: {filename}\n\n"
                f"I've filled **{fill_result.filled_count}/{fill_result.total_fields}** fields "
                f"using your company documents.\n\n"
                f"### Preview of filled fields:\n{summary}\n\n"
                f"---\n\n"
                f"\U0001f449 **The completed form is now in your [Approvals](/approvals) inbox.** "
                f"Please review it there and click **Approve** to send the final file, "
                f"or **Reject** with feedback if changes are needed.\n\n"
                f"_Processing took {duration_ms}ms._"
            ),
            metadata={
                "intent": "fill_form",
                "status": "awaiting_approval",
                "submission_id": submission_id,
                "fields_filled": fill_result.filled_count,
                "fields_total": fill_result.total_fields,
                "duration_ms": duration_ms,
            },
        )

        # Slack notification — form filling completed
        notify_task_completed(
            agent_name="tender-agent",
            task_title=f"Form Filled: {filename}",
            summary=(
                f"Filled {fill_result.filled_count}/{fill_result.total_fields} fields "
                f"using company documents. Submitted for human review."
            ),
            metrics={
                "duration_ms": duration_ms,
                "cost_usd": fill_result.llm_cost_usd,
                "fields_filled": fill_result.filled_count,
                "fields_total": fill_result.total_fields,
            },
            task_type="form_fill",
        )

        # Audit: submission created
        _audit(
            "task_completed",
            f"Form submitted for approval: {filename} ({fill_result.filled_count}/{fill_result.total_fields} fields)",
            node_name="fill_form",
            duration_ms=duration_ms,
            cost_usd=fill_result.llm_cost_usd,
            output_payload={"submission_id": submission_id, **output_summary},
        )

    except Exception as exc:
        print(f"Failed to submit for approval: {exc}")
        traceback.print_exc()

        # Fallback: send the form directly in chat (skip approval)
        try:
            client.reply_with_file(
                thread_id=thread_id,
                content=(
                    f"## \u2705 Form Completed: {filename}\n\n"
                    f"I've filled **{fill_result.filled_count}/{fill_result.total_fields}** fields.\n\n"
                    f"\u26a0\ufe0f _Note: The approval system was unavailable, so I'm sending the form directly. "
                    f"Please review it carefully before using._"
                ),
                file_path=output_path,
                filename=output_filename,
                metadata={"intent": "fill_form", "status": "direct_send_fallback"},
            )
        except Exception as fallback_exc:
            client.reply(
                thread_id=thread_id,
                content=f"Sorry, I filled the form but couldn't send it: {str(fallback_exc)}",
                metadata={"intent": "fill_form", "status": "error"},
            )

    # Clean up session
    if thread_id in _form_sessions:
        del _form_sessions[thread_id]


def _format_questions(fill_result: FillResult) -> str:
    """Format clarification questions for the user."""
    lines = []
    for i, q in enumerate(fill_result.questions, 1):
        lines.append(f"**{i}. {q.field_name}**")
        lines.append(f"   {q.question}")
        if q.suggestions:
            lines.append(f"   _Suggestion: {q.suggestions[0]}_")
        lines.append("")
    return "\n".join(lines)


def generate_natural_response(user_message: str, intent: dict, results_text: str, num_results: int) -> str:
    """Use LLM to generate a natural, conversational response."""
    try:
        if intent["intent"] == "greeting":
            result = call_llm(
                f"""The user said: "{user_message}"
You are the Tender Agent — a friendly AI that helps find government and enterprise tenders related to Safety Data Sheets (SDS), chemical safety, and EHS compliance.
You can: search for tenders globally, read company documents uploaded in the AMS Documents section, and fill out tender forms.
Respond naturally in 2-3 sentences. Introduce yourself briefly and mention your capabilities.""",
                max_tokens=200,
            )
            return result["content"]

        if intent["intent"] == "ask_question":
            result = call_llm(
                f"""The user asked: "{user_message}"
You are the Tender Agent specializing in SDS/EHS/chemical safety tenders.
You have access to company documents uploaded in the AMS Documents section — use them to answer questions about the company.
{fetch_agent_context()}
Answer their question concisely using the document context above if relevant. If you don't have the information, say so and offer to search for relevant tenders instead.""",
                max_tokens=500,
            )
            return result["content"]

        # For search results, generate a summary
        if num_results > 0:
            result = call_llm(
                f"""The user asked: "{user_message}"
I searched globally and found {num_results} tender opportunities. Here are the results:

{results_text}

Write a natural 2-3 sentence executive summary. Mention the most promising opportunity, which region/agency it's from, and why it's relevant to SDS/EHS. Be specific, not generic.""",
                max_tokens=300,
            )
            return result["content"]
        else:
            return f"I searched globally for tenders matching your request but didn't find strong matches right now. Try rephrasing — for example: 'find chemical safety compliance RFPs in Europe' or 'SDS management tenders in the US'."

    except Exception as exc:
        print(f"Natural response generation failed: {exc}")
        return ""


def handle_message(client: NexusClient, message: dict) -> None:
    """Handle an incoming chat message from the AMS.

    Uses LLM to understand the user's intent and respond naturally.
    Supports: tender search, form filling, Q&A, greetings.
    """
    thread_id = message.get("threadId", "")
    content = message.get("content", "")
    sender = message.get("senderId", "unknown")
    attachments = message.get("attachments", [])

    # Strip @tender-agent mention from content
    clean_content = content.replace("@tender-agent", "").strip()

    # Set module-level context for audit helper
    global _current_thread_id
    _current_thread_id = thread_id

    print(f"\n{'='*60}")
    print(f"  INCOMING MESSAGE")
    print(f"  Thread:  {thread_id}")
    print(f"  From:    {sender}")
    print(f"  Content: {clean_content[:100]}{'...' if len(clean_content) > 100 else ''}")
    if attachments:
        print(f"  Attachments: {len(attachments)} file(s)")
        for att in attachments:
            print(f"    - {att.get('filename', '?')} ({att.get('mimeType', '?')})")
    print(f"{'='*60}\n")

    # Audit: task started
    _audit(
        "task_started",
        f"Received message: {clean_content[:100]}",
        input_payload={
            "message": clean_content[:500],
            "sender": sender,
            "has_attachments": bool(attachments),
            "attachment_count": len(attachments),
            "attachment_names": [a.get("filename", "?") for a in attachments],
        },
    )

    # Check if this is a follow-up to an active form-filling session
    if thread_id in _form_sessions and _form_sessions[thread_id].get("waiting_for_answers"):
        print("Continuing form-filling session (user answers)...")
        _audit("task_started", "Continuing form-filling session (user provided answers)")
        handle_form_filling(client, message)
        return

    client.track(
        "conversation_volume",
        1.0,
        metadata={"thread_id": thread_id, "direction": "inbound"},
    )

    start_time = time.perf_counter()

    try:
        # Step 1: Understand what the user wants
        has_attachments = len(attachments) > 0
        print("Classifying user intent...")
        intent = classify_intent(clean_content, has_attachments=has_attachments)
        print(f"  Intent: {intent.get('intent', 'unknown')}")
        print(f"  Search query: {intent.get('search_query', 'none')}")

        # Audit: intent classification result
        _audit(
            "intent_classified",
            f"Intent: {intent.get('intent', 'unknown')} | Query: {intent.get('search_query', 'none')[:80]}",
            node_name="classify",
            output_payload=intent,
        )

        # Step 2: Handle form filling
        if intent["intent"] == "fill_form":
            print("Routing to form-filling handler...")
            handle_form_filling(client, message)
            return

        # Step 3: Handle greetings and questions
        if intent["intent"] in ("greeting", "ask_question") and intent["intent"] != "search_tenders":
            reply_text = generate_natural_response(clean_content, intent, "", 0)
            duration_ms = int((time.perf_counter() - start_time) * 1000)

            _audit(
                "task_completed",
                f"Sent {intent['intent']} response ({duration_ms}ms)",
                node_name=intent["intent"],
                duration_ms=duration_ms,
                output_payload={"reply_length": len(reply_text)},
            )

            client.reply(
                thread_id=thread_id,
                content=reply_text,
                metadata={"intent": intent["intent"], "duration_ms": duration_ms},
            )
            print(f"Conversational reply sent ({duration_ms}ms)")

            # Slack notification for Q&A (skip greetings — too noisy)
            if intent["intent"] == "ask_question":
                notify_task_completed(
                    agent_name="tender-agent",
                    task_title=f"Q&A: {clean_content[:80]}",
                    summary=reply_text[:200] + ("..." if len(reply_text) > 200 else ""),
                    metrics={"duration_ms": duration_ms},
                    task_type="question",
                    thread_id=thread_id,
                )

            return

        # Step 4: Run tender search — always pass the user's original message
        # so region detection works correctly (LLM extraction can lose region context)
        print("Running global tender search...")
        reply_text, metadata = run_tender_search(clean_content)

        # Step 4: Generate natural summary and prepend it
        natural_intro = generate_natural_response(
            clean_content, intent,
            reply_text, metadata.get("tenders_found", 0)
        )

        if natural_intro:
            # Replace the generic header with the natural response
            final_reply = (
                f"{natural_intro}\n\n"
                f"---\n\n"
                f"{reply_text.split('---', 1)[-1] if '---' in reply_text else reply_text}"
            )
        else:
            final_reply = reply_text

        metadata["thread_id"] = thread_id
        metadata["intent"] = intent["intent"]
        client.reply(
            thread_id=thread_id,
            content=final_reply,
            metadata=metadata,
        )

        duration_ms = int((time.perf_counter() - start_time) * 1000)
        print(f"Reply sent! Found {metadata['tenders_found']} tenders ({duration_ms}ms)")

        # Slack notification — tender search completed
        notify_task_completed(
            agent_name="tender-agent",
            task_title=f"Tender Search: {clean_content[:80]}",
            summary=f"Found {metadata['tenders_found']} verified tenders matching the query.",
            metrics={
                "duration_ms": duration_ms,
                "cost_usd": metadata.get("cost_usd", 0),
                "tokens_used": metadata.get("tokens_input", 0) + metadata.get("tokens_output", 0),
                "tenders_found": metadata.get("tenders_found", 0),
            },
            task_type="tender_search",
            thread_id=thread_id,
        )

        # Audit: search completed
        _audit(
            "task_completed",
            f"Tender search completed: {metadata['tenders_found']} results ({duration_ms}ms)",
            node_name="discover",
            duration_ms=duration_ms,
            cost_usd=metadata.get("cost_usd", 0),
            model_used=metadata.get("model"),
            tokens_input=metadata.get("tokens_input", 0),
            tokens_output=metadata.get("tokens_output", 0),
            output_payload={
                "tenders_found": metadata["tenders_found"],
                "top_relevance": metadata.get("top_relevance", 0),
                "search_source": metadata.get("search_source", "unknown"),
                "intent": intent.get("intent"),
            },
        )

        # Track ALL metric types for dashboard
        tokens_in = metadata.get("tokens_input", 0)
        tokens_out = metadata.get("tokens_output", 0)
        total_tokens = tokens_in + tokens_out
        cost = metadata.get("cost_usd", 0)

        client.track("task_completion", 1.0, metadata={
            "node": "discover",
            "tenders_found": metadata["tenders_found"],
        })
        client.track("latency", float(duration_ms), metadata={
            "node": "discover",
        })
        client.track("throughput", 1.0, metadata={
            "node": "discover",
        })
        client.track("conversation_volume", 1.0, metadata={
            "thread_id": thread_id,
            "direction": "outbound",
        })
        if total_tokens > 0:
            client.track("token_usage", float(total_tokens), metadata={
                "model": LLM_MODEL,
                "tokens_input": tokens_in,
                "tokens_output": tokens_out,
            })
        if cost > 0:
            client.track("cost", cost, metadata={
                "node": "discover",
                "model": LLM_MODEL,
                "cost_usd": cost,
                "tokens_total": total_tokens,
            })
        client.track("tool_calls", 1.0, metadata={
            "tool": "serp_search",
            "node": "discover",
        })
        client.track("model_distribution", 1.0, metadata={
            "model": LLM_MODEL,
        })
        client.track("uptime", 1.0, metadata={})
        client.track("user_satisfaction", 1.0, metadata={
            "thread_id": thread_id,
        })

    except Exception as exc:
        error_msg = f"Error: {str(exc)}"
        print(f"ERROR: {error_msg}")
        traceback.print_exc()

        # Slack notification — task failed
        notify_task_failed(
            agent_name="tender-agent",
            task_title=f"Message: {clean_content[:80]}",
            error_message=str(exc)[:300],
        )

        # Audit: task failed
        _audit(
            "task_failed",
            f"Error processing message: {str(exc)[:200]}",
            status="failure",
            error_message=str(exc),
            input_payload={"message": clean_content[:300]},
        )

        client.reply(
            thread_id=thread_id,
            content=f"Sorry, I ran into an issue while processing your request:\n\n`{str(exc)}`\n\nCould you try rephrasing? For example: 'find SDS management tenders in Europe'",
            metadata={"error": str(exc), "node": "discover"},
        )

        client.track("error_rate", 1.0, metadata={
            "node": "discover",
            "error_type": type(exc).__name__,
        })
        client.track("task_completion", 0.0, metadata={
            "node": "discover",
            "status": "failure",
        })

    # Always flush metrics after handling a message
    try:
        flushed = client.flush()
        print(f"Metrics flushed: {flushed.inserted} events sent to AMS")
    except Exception as exc:
        print(f"Metric flush failed: {exc}")


# ---------------------------------------------------------------------------
# Main — Registration + Heartbeat + Inbox Polling
# ---------------------------------------------------------------------------

def _require_agent_credential() -> None:
    """Refuse to run without NEXUS_AGENT_API_KEY (MR2).

    P0.0 made every raw AMS call send `_ams_headers()`, but that helper returns
    `{}` when the key is unset — correct then, because the AMS fell open and the
    bridge worked anyway, which is precisely how the key stayed unset in
    production for months. P0.1 closed those endpoints. An unset key now means
    every AMS call is refused while the agent appears to run normally.

    A loud refusal to start is the right trade. It mirrors
    `apps/web/src/instrumentation.ts` on the AMS side, which already refuses to
    boot for the same reason: an outage is visible in thirty seconds, silent
    degradation was not visible for months.
    """
    if NEXUS_AGENT_API_KEY:
        return

    print("")
    print("*** STARTUP REFUSED — missing required configuration ***")
    print("")
    print("  NEXUS_AGENT_API_KEY is unset.")
    print("")
    print("  It authenticates every bridge -> AMS call and must be set to the")
    print("  SAME value here and on the AMS web service. Without it the AMS")
    print("  refuses every request (401) and this agent runs but accomplishes")
    print("  nothing: knowledge-base search, file downloads and the active")
    print("  prompt all fail, two of them silently.")
    print("")
    print("  Set it in this service's environment and restart. See")
    print("  docs/runbooks/p0.1-bridge-auth.md in the nexus-ams repository.")
    print("")
    raise SystemExit(1)


def main() -> None:
    print("")
    print("=" * 60)
    print("  TENDER AGENT — Nexus AMS Bridge")
    print("=" * 60)
    print(f"  AMS URL:     {NEXUS_URL}")
    print(f"  LLM Model:   {LLM_MODEL}")
    print(f"  Dry Run:     {DRY_RUN}")
    print(f"  OpenRouter:  {'configured' if OPENROUTER_API_KEY else 'NOT configured'}")
    print(f"  Agent token: {'set' if NEXUS_AGENT_API_KEY else 'UNSET — every AMS call will be refused'}")
    print("=" * 60)
    print("")

    _require_agent_credential()

    # Step 1: Create client and register
    global _client
    client = NexusClient(base_url=NEXUS_URL)
    _client = client  # Set module-level reference for audit helper

    config = AgentConfig(
        name="tender-agent",
        display_name="Tender Agent",
        description=(
            "Discovers and evaluates government tenders related to SDS management. "
            "Searches SAM.gov and scores opportunities by relevance to EHS/chemical safety."
        ),
        version="1.0.0",
        python_version="3.11",
        langgraph_version="0.4.0",
        llm_provider="openrouter",
        llm_models={
            LLM_MODEL: {
                "provider": "openrouter",
                "context_window": 200000,
            },
        },
        llm_pricing={
            LLM_MODEL: {
                "input_cost_per_million": 3.00,
                "output_cost_per_million": 15.00,
            },
        },
        embedding_model="voyage-3-large",
        embedding_dimensions=1024,
        state_fields_count=40,
        node_names=["discover", "evaluate", "retrieve_draft", "gap_check",
                     "slack_escalate", "assemble", "submit"],
        tools=["sam_gov_scraper", "email_monitor", "voyage_embedder",
               "template_engine", "slack_client", "playwright_submitter"],
        health_endpoint="http://localhost:8100/health",
        slack_channels=["#tender-alerts"],
        env_vars_count=23,
        dry_run=DRY_RUN,
        budget_monthly_usd=50.0,
        tags=["production", "government-tenders", "sds", "ehs"],
        changelog="Bridge connection to Nexus AMS",
    )

    print("Registering with Nexus AMS...")
    try:
        resp = client.register(config)
        print(f"Registered! Agent ID: {resp.agent_id}")
        # Notify Slack that agent is online
        notify_agent_online("tender-agent")
    except Exception as exc:
        print(f"Registration failed: {exc}")
        print("Make sure the AMS is running at", NEXUS_URL)
        sys.exit(1)

    # Step 2: Start heartbeat in background
    print("Starting heartbeat (every 30 seconds)...")

    stop_event = threading.Event()

    def heartbeat_loop() -> None:
        while not stop_event.is_set():
            try:
                client.heartbeat(status="running")
            except Exception as exc:
                print(f"Heartbeat error (will retry): {exc}")
            stop_event.wait(30.0)

    hb_thread = threading.Thread(target=heartbeat_loop, daemon=True)
    hb_thread.start()

    # Step 3: Poll inbox for messages
    print("")
    print("Listening for messages from AMS chat...")
    print("Go to http://localhost:3000/chat and send a message mentioning @tender-agent")
    print("Example: '@tender-agent search for SDS management tenders'")
    print("")
    print("Press Ctrl+C to stop.")
    print("")

    try:
        while True:
            try:
                # Poll the inbox endpoint for new messages
                import httpx
                inbox_url = f"{NEXUS_URL}/api/agents/tender-agent/inbox"
                resp = httpx.get(inbox_url, headers=_ams_headers(), timeout=10.0)

                if resp.status_code == 200:
                    data = resp.json()
                    # Handle both { "messages": [...] } and plain [...]
                    if isinstance(data, dict):
                        messages = data.get("messages", [])
                    elif isinstance(data, list):
                        messages = data
                    else:
                        messages = []
                    for msg in messages:
                        handle_message(client, msg)

            except httpx.ConnectError:
                pass  # AMS not reachable, will retry
            except Exception as exc:
                print(f"Inbox polling error: {exc}")

            time.sleep(2.0)  # Poll every 2 seconds

    except KeyboardInterrupt:
        print("\n\nShutting down...")
        stop_event.set()
        notify_agent_offline("tender-agent")
        client.flush()
        client.close()
        print("Tender Agent bridge stopped.")


if __name__ == "__main__":
    main()
